from flask import Flask, render_template, request, send_file
from googleapiclient.discovery import build
from isodate import parse_duration
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__, template_folder="templates")
YOUTUBE_API_KEY = "AIzaSyB7FKZ5tP8bbh6TaixauE0y551D1-j7qEM"  # Replace with your real API key

# ------------- YouTube API -------------
def get_playlist_videos(playlist_url):
    youtube = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)

    if "list=" not in playlist_url:
        return []
    playlist_id = playlist_url.split("list=")[-1].split("&")[0]

    videos = []
    next_page_token = None

    while True:
        pl_request = youtube.playlistItems().list(
            part="snippet",
            playlistId=playlist_id,
            maxResults=50,
            pageToken=next_page_token
        )
        pl_response = pl_request.execute()

        video_ids = [item['snippet']['resourceId']['videoId'] for item in pl_response['items']]
        if not video_ids:
            break

        vid_request = youtube.videos().list(
            part="contentDetails,snippet",
            id=",".join(video_ids)
        )
        vid_response = vid_request.execute()

        for item in vid_response["items"]:
            title = item["snippet"]["title"]
            duration = parse_duration(item["contentDetails"]["duration"]).total_seconds() / 60
            url = f"https://www.youtube.com/watch?v={item['id']}"
            videos.append((title, round(duration), url))

        next_page_token = pl_response.get("nextPageToken")
        if not next_page_token:
            break

    return videos

# ------------- Schedulers -------------
def generate_schedule_by_duration(videos, daily_duration):
    schedule = []
    day = []
    time_used = 0

    for title, duration, url in videos:
        if time_used + duration > daily_duration:
            schedule.append(day)
            day = []
            time_used = 0
        day.append((title, duration, url))
        time_used += duration

    if day:
        schedule.append(day)
    return schedule

def generate_schedule_by_days(videos, num_days):
    total_duration = sum(duration for _, duration, _ in videos)
    if num_days <= 0 or total_duration == 0:
        return [[] for _ in range(num_days)]

    per_day_target = total_duration / num_days
    schedule = [[] for _ in range(num_days)]

    day_index = 0
    time_left_in_day = per_day_target

    for title, duration, url in videos:
        while duration > 0:
            if duration <= time_left_in_day:
                schedule[day_index].append((title, round(duration, 2), url))
                time_left_in_day -= duration
                duration = 0
            else:
                schedule[day_index].append((f"{title} (Part)", round(time_left_in_day, 2), url))
                duration -= time_left_in_day
                day_index += 1
                if day_index >= num_days:
                    day_index = num_days - 1
                time_left_in_day = per_day_target

        if time_left_in_day <= 0.01 and day_index < num_days - 1:
            day_index += 1
            time_left_in_day = per_day_target

    return schedule

# ------------- Routes -------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/schedule', methods=['POST'])
def show_schedule():
    playlist_url = request.form['playlist_url']
    mode = request.form['mode']
    videos = get_playlist_videos(playlist_url)

    if mode == 'duration':
        daily_duration = int(request.form['daily_duration'])
        schedule = generate_schedule_by_duration(videos, daily_duration)
    else:
        num_days = int(request.form['num_days'])
        schedule = generate_schedule_by_days(videos, num_days)

    global last_schedule
    last_schedule = schedule
    return render_template('schedule.html', schedule=schedule)

@app.route('/download/excel')
def download_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Watch Schedule"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alternate_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    # Write headers
    headers = ['Day', 'Title', 'Duration (min)', 'Total (min)', 'URL']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Fill in video data with alternating row colors
    row = 2
    for day_num, day in enumerate(last_schedule, start=1):
        start_row = row
        day_total = sum(duration for _, duration, _ in day)

        for title, duration, url in day:
            ws.cell(row=row, column=1, value=f"Day {day_num}").alignment = center_align
            ws.cell(row=row, column=2, value=title).alignment = center_align
            ws.cell(row=row, column=3, value=duration).alignment = center_align
            ws.cell(row=row, column=5, value=url).alignment = center_align

            # Apply alternate row fill
            if row % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = alternate_fill
            row += 1

        # Merge the "Total" cell for this day's rows and set the total value
        ws.merge_cells(start_row=start_row, start_column=4, end_row=row - 1, end_column=4)
        total_cell = ws.cell(row=start_row, column=4, value=day_total)
        total_cell.alignment = center_align

        # Apply fill and font to merged cell if it lands on alternate row
        if start_row % 2 == 0:
            total_cell.fill = alternate_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    # Return as downloadable file
    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(mem, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='schedule.xlsx', as_attachment=True)

if __name__ == '__main__':
    last_schedule = []
    app.run()

